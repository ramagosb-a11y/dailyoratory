from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "output"
OUTPUT_PATH = OUTPUT_DIR / "Daily Oratory Media Library Admin - dropdowns.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="7A2533")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REQUIRED_FILL = PatternFill("solid", fgColor="F4D7A1")
BODY_FILL = PatternFill("solid", fgColor="FFFDF7")
TITLE_FILL = PatternFill("solid", fgColor="0D2038")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=18)
SUBTITLE_FILL = PatternFill("solid", fgColor="F3EAD8")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9C9AA"),
    right=Side(style="thin", color="D9C9AA"),
    top=Side(style="thin", color="D9C9AA"),
    bottom=Side(style="thin", color="D9C9AA"),
)

MEDIA_HEADERS = [
    "Media ID",
    "Title",
    "Slug",
    "Description",
    "Short Description",
    "Media Type",
    "Category",
    "Tags",
    "Topic",
    "Audience",
    "YouTube URL",
    "YouTube Video ID",
    "YouTube Playlist ID",
    "Google Slides URL",
    "Google Slides Embed URL",
    "Google Drive File URL",
    "Google Drive File ID",
    "Image URL",
    "Thumbnail URL",
    "Alt Text",
    "Speaker / Creator",
    "Source Name",
    "Source URL",
    "Copyright Status",
    "License Notes",
    "Featured",
    "Collection ID",
    "Related Daily Oratory Links",
    "Related Scripture References",
    "Related Catechism References",
    "Status",
    "Sort Order",
    "Published Date",
    "Updated Date",
    "Notes",
]

REQUIRED_MEDIA_HEADERS = {
    "Media ID",
    "Title",
    "Slug",
    "Description",
    "Short Description",
    "Media Type",
    "Category",
    "Topic",
    "Audience",
    "Alt Text",
    "Speaker / Creator",
    "Source Name",
    "Copyright Status",
    "Featured",
    "Status",
    "Sort Order",
    "Published Date",
    "Updated Date",
}

STATUS_OPTIONS = ["Draft", "Review", "Approved", "Hidden", "Do Not Publish"]
COPYRIGHT_OPTIONS = [
    "original",
    "owned-by-daily-oratory",
    "permission-granted",
    "public-domain",
    "creative-commons",
    "external-embed",
    "needs-review",
    "do-not-publish",
]
MEDIA_TYPE_OPTIONS = [
    "youtube-video",
    "youtube-playlist",
    "google-slides",
    "google-drive-image",
    "google-drive-file",
    "external-image",
    "pdf",
    "collection",
]
CATEGORY_ROWS = [
    ("category-prayer", "Prayer", "prayer", "Prayer resources for silence, petitions, devotions, and conversation with God.", "hands-praying", 10, "Approved"),
    ("category-scripture", "Scripture", "scripture", "Resources rooted in the Word of God and Catholic interpretation.", "book-open", 20, "Approved"),
    ("category-mass", "Mass", "mass", "Formation resources for the Holy Mass and Catholic worship.", "church", 30, "Approved"),
    ("category-sacraments", "Sacraments", "sacraments", "Resources explaining sacramental life and grace.", "sparkles", 40, "Approved"),
    ("category-confession", "Confession", "confession", "Mercy, repentance, and reconciliation resources.", "heart", 50, "Approved"),
    ("category-adoration", "Adoration", "adoration", "Eucharistic prayer resources and reverent formation.", "sun", 60, "Approved"),
    ("category-rosary", "Rosary", "rosary", "Marian prayer resources and mystery meditation guides.", "circle", 70, "Approved"),
    ("category-saints", "Saints", "saints", "Saint stories, prayer cards, and holiness companions.", "star", 80, "Approved"),
    ("category-devotions", "Devotions", "devotions", "Traditional devotions ordered toward Christ.", "candles", 90, "Approved"),
    ("category-formation", "Formation", "formation", "Catholic doctrine, prayer, virtue, and discipleship resources.", "graduation-cap", 100, "Approved"),
    ("category-ocia", "OCIA / Becoming Catholic", "ocia", "Resources for seekers, catechumens, candidates, and inquirers.", "footprints", 110, "Approved"),
    ("category-family", "Family and Domestic Church", "family", "Prayer and formation resources for family life.", "home", 120, "Approved"),
    ("category-liturgical-living", "Liturgical Living", "liturgical-living", "Resources for living the Church year in daily life.", "calendar-days", 130, "Approved"),
    ("category-church-history", "Church History", "church-history", "Historical resources about the Church across the centuries.", "landmark", 140, "Approved"),
    ("category-church-fathers", "Church Fathers", "church-fathers", "Early Christian teaching resources and source guides.", "scroll-text", 150, "Approved"),
    ("category-indulgences", "Indulgences", "indulgences", "Resources for understanding indulgences and their conditions.", "award", 160, "Approved"),
    ("category-youth", "Youth / Students", "youth-students", "Resources for young Catholics and school formation.", "notebook-pen", 170, "Approved"),
    ("category-retreats", "Retreats", "retreats", "Prayer and formation resources for retreat settings.", "trees", 180, "Approved"),
    ("category-talks", "Homilies / Talks", "homilies-talks", "Talks, teaching sessions, and spoken formation resources.", "mic", 190, "Approved"),
    ("category-slides", "Slides and Teaching Resources", "slides-teaching-resources", "Slides, presentations, and teaching materials.", "presentation", 200, "Approved"),
    ("category-images", "Images and Prayer Cards", "images-prayer-cards", "Visual resources, prayer cards, and sacred images.", "image", 210, "Approved"),
]
COLLECTION_ROWS = [
    ("collection-start-here-prayer", "Start Here for Prayer", "start-here-for-prayer", "A gentle first set of resources for beginning prayer with Daily Oratory.", "media-welcome-daily-oratory, media-prayer-card-image", "TRUE", 10, "Approved"),
    ("collection-first-time-at-mass", "First Time at Mass", "first-time-at-mass", "Helpful resources for those attending Mass for the first time or returning after time away.", "media-welcome-daily-oratory", "TRUE", 20, "Approved"),
    ("collection-eucharistic-adoration", "Eucharistic Adoration", "eucharistic-adoration", "Prayerful media to support Eucharistic faith and reverence.", "media-prayer-card-image", "TRUE", 30, "Approved"),
    ("collection-ocia-exploring-faith", "OCIA and Exploring the Faith", "ocia-and-exploring-the-faith", "Clear and welcoming resources for those learning Catholicism step by step.", "media-catholic-formation-slides", "TRUE", 40, "Approved"),
    ("collection-family-prayer-home", "Family Prayer at Home", "family-prayer-at-home", "A simple starter collection for families building a Catholic prayer rhythm.", "media-prayer-card-image", "TRUE", 50, "Approved"),
    ("collection-mass-explained", "The Mass Explained", "the-mass-explained", "Resources that help viewers understand Catholic worship and the Eucharist.", "media-welcome-daily-oratory", "TRUE", 60, "Approved"),
]
FEATURED_OPTIONS = ["TRUE", "FALSE"]


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    media_items = wb.create_sheet("Media_Items")
    categories = wb.create_sheet("Media_Categories")
    collections = wb.create_sheet("Featured_Collections")
    settings = wb.create_sheet("Settings")
    change_log = wb.create_sheet("Change_Log")
    copyright_review = wb.create_sheet("Copyright_Review")
    lists = wb.create_sheet("Lists")

    build_lists_sheet(lists)
    build_instructions_sheet(instructions)
    build_media_items_sheet(media_items)
    build_categories_sheet(categories)
    build_collections_sheet(collections)
    build_settings_sheet(settings)
    build_change_log_sheet(change_log)
    build_copyright_sheet(copyright_review)

    lists.sheet_state = "hidden"
    wb.save(OUTPUT_PATH)
    print(f"Saved workbook to {OUTPUT_PATH}")


def build_instructions_sheet(ws):
    ws.merge_cells("A1:F1")
    ws["A1"] = "Daily Oratory Media Library Admin"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "Use this workbook as your upload-ready Google Sheets admin template. Gold headers in Media_Items are required. Dropdown cells are built into the key fields."
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].alignment = Alignment(wrap_text=True)

    rows = [
        ("Media_Items", "Main library records. Required columns are highlighted in gold."),
        ("Media_Categories", "Approved category list used for browsing and filters."),
        ("Featured_Collections", "Grouped media sets such as Start Here for Prayer."),
        ("Settings", "Publishing rules and future integration notes."),
        ("Change_Log", "Optional manual log for editorial changes."),
        ("Copyright_Review", "Track approval before public publishing."),
        ("", ""),
        ("Dropdown fields in Media_Items", "Media Type, Category, Copyright Status, Featured, Collection ID, Status"),
        ("Required fields", "Look for gold headers with a Required comment."),
        ("Status values", ", ".join(STATUS_OPTIONS)),
        ("Copyright values", ", ".join(COPYRIGHT_OPTIONS)),
        ("Media type values", ", ".join(MEDIA_TYPE_OPTIONS)),
        ("Publishing rule", "Only Approved rows with cleared copyright and valid public media links should appear publicly."),
        ("Drive rule", "Only use files intentionally shared as Anyone with the link can view."),
    ]
    ws["A4"] = "Section"
    ws["B4"] = "Notes"
    style_header_cell(ws["A4"])
    style_header_cell(ws["B4"])

    for idx, (left, right) in enumerate(rows, start=5):
        ws[f"A{idx}"] = left
        ws[f"B{idx}"] = right
        style_body_row(ws, idx, 1, 2)

    set_column_widths(ws, {
        "A": 28,
        "B": 90,
        "C": 4,
        "D": 24,
        "E": 64,
    })
    ws.freeze_panes = "A4"


def build_media_items_sheet(ws):
    for col_idx, header in enumerate(MEDIA_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell)
        if header in REQUIRED_MEDIA_HEADERS:
            cell.fill = REQUIRED_FILL
            cell.font = Font(color="0D2038", bold=True)
            cell.comment = Comment("Required field", "Daily Oratory")

    sample_rows = [
        [
            "media-welcome-daily-oratory", "Welcome to Daily Oratory", "welcome-to-daily-oratory",
            "TODO: add a full original description for the first public media item.",
            "TODO: add a short description.", "youtube-video", "formation", "Prayer, Formation, Welcome",
            "Introduction", "Catholics, returning Catholics, people exploring the Catholic faith",
            "TODO: paste a public YouTube URL", "", "", "", "", "", "", "", "",
            "TODO: add clear alt text if a thumbnail or image is used.", "Daily Oratory", "YouTube", "",
            "needs-review", "Replace after copyright review.", "FALSE", "collection-start-here-prayer",
            "/pray, /explore, /formation", "", "", "Draft", 10, "2026-05-17", "2026-05-17",
            "Placeholder starter record. Do not publish until the real public URL is added.",
        ],
        [
            "media-catholic-formation-slides", "Catholic Formation Slides", "catholic-formation-slides",
            "TODO: describe the purpose of this slide deck for parish, OCIA, or family formation.",
            "Draft placeholder for a future Google Slides presentation.", "google-slides",
            "slides-teaching-resources", "Slides, Formation, Teaching", "Catholic Formation Slides",
            "Catholics, OCIA users, parish groups, formation leaders", "", "", "",
            "TODO: paste the Google Slides viewer URL", "TODO: paste the Publish to web embed URL", "", "", "", "",
            "Slide deck preview for Catholic formation.", "Daily Oratory", "Google Slides", "",
            "needs-review", "Only publish after the deck is intentionally public.", "FALSE",
            "collection-ocia-exploring-faith", "/formation, /ocia, /explore", "", "", "Draft", 20,
            "2026-05-17", "2026-05-17", "Use Publish to web > Embed for best on-site display.",
        ],
        [
            "media-prayer-card-image", "Prayer Card Image", "prayer-card-image",
            "TODO: describe the image, prayer card, or visual resource being displayed.",
            "Draft placeholder for a public prayer image.", "google-drive-image",
            "images-prayer-cards", "Image, Prayer Card, Formation", "Prayer Image",
            "Catholics, families, parish groups", "", "", "", "", "",
            "TODO: paste the public Google Drive file URL", "", "", "",
            "TODO: add alt text that clearly describes the image.", "Daily Oratory", "Google Drive", "",
            "needs-review", "Only publish intentionally public images.", "FALSE",
            "collection-family-prayer-home", "/family, /pray, /saints", "", "", "Draft", 30,
            "2026-05-17", "2026-05-17", "Set sharing to Anyone with the link can view before approval.",
        ],
    ]
    for row_idx, row in enumerate(sample_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_body_row(ws, row_idx, 1, len(MEDIA_HEADERS))

    apply_media_validations(ws)
    set_column_widths(ws, {
        "A": 24, "B": 28, "C": 24, "D": 42, "E": 28, "F": 18, "G": 20, "H": 22, "I": 20, "J": 28,
        "K": 28, "L": 18, "M": 18, "N": 28, "O": 30, "P": 28, "Q": 20, "R": 22, "S": 22, "T": 28,
        "U": 22, "V": 18, "W": 22, "X": 22, "Y": 24, "Z": 12, "AA": 28, "AB": 28, "AC": 24, "AD": 24,
        "AE": 16, "AF": 12, "AG": 16, "AH": 16, "AI": 34,
    })
    ws.freeze_panes = "A2"


def build_categories_sheet(ws):
    headers = ["Category ID", "Title", "Slug", "Description", "Icon Name", "Sort Order", "Status"]
    add_header_row(ws, headers)
    for row_idx, row in enumerate(CATEGORY_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_body_row(ws, row_idx, 1, len(headers))
    add_dropdown(ws, "G2:G500", '=Lists!$D$2:$D$6')
    set_column_widths(ws, {"A": 22, "B": 28, "C": 22, "D": 50, "E": 18, "F": 12, "G": 16})
    ws.freeze_panes = "A2"


def build_collections_sheet(ws):
    headers = ["Collection ID", "Title", "Slug", "Description", "Media IDs", "Featured", "Sort Order", "Status"]
    add_header_row(ws, headers)
    for row_idx, row in enumerate(COLLECTION_ROWS, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_body_row(ws, row_idx, 1, len(headers))
    add_dropdown(ws, "F2:F500", '=Lists!$E$2:$E$3')
    add_dropdown(ws, "H2:H500", '=Lists!$D$2:$D$6')
    set_column_widths(ws, {"A": 28, "B": 28, "C": 24, "D": 54, "E": 34, "F": 12, "G": 12, "H": 16})
    ws.freeze_panes = "A2"


def build_settings_sheet(ws):
    headers = ["Setting Key", "Setting Value", "Notes"]
    rows = [
        ("Sheet Name", "Daily Oratory Media Library Admin", "Recommended Google Sheet name after upload."),
        ("Publishing Rule", "Only Approved items with cleared copyright may appear publicly.", "Draft, Review, Hidden, and Do Not Publish must stay off the site."),
        ("YouTube Embed", "Use youtube-nocookie embeds", "Do not autoplay with sound."),
        ("Google Slides Embed", "Use Publish to web embed URL if possible", "Fallback is Open Slides button."),
        ("Google Drive Preview", "Use public file preview links only", "Only Anyone with the link can view files should be added."),
        ("Future Feed Option", "Google Sheet export or Apps Script JSON endpoint", "Current site version uses local data first."),
    ]
    add_header_row(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_body_row(ws, row_idx, 1, len(headers))
    set_column_widths(ws, {"A": 22, "B": 38, "C": 56})
    ws.freeze_panes = "A2"


def build_change_log_sheet(ws):
    headers = ["Timestamp", "Editor", "Change Type", "Media ID", "Notes"]
    add_header_row(ws, headers)
    row = ["2026-05-17 16:30", "Owner", "Template created", "media-welcome-daily-oratory", "Initial starter workbook with dropdowns and required-field marking."]
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=2, column=col_idx, value=value)
    style_body_row(ws, 2, 1, len(headers))
    set_column_widths(ws, {"A": 20, "B": 18, "C": 22, "D": 24, "E": 56})
    ws.freeze_panes = "A2"


def build_copyright_sheet(ws):
    headers = ["Media ID", "Title", "Source", "Copyright Status", "Reviewed By", "Review Date", "Notes"]
    add_header_row(ws, headers)
    rows = [
        ("media-welcome-daily-oratory", "Welcome to Daily Oratory", "YouTube", "needs-review", "", "", "Fill this in before publishing the item."),
        ("media-catholic-formation-slides", "Catholic Formation Slides", "Google Slides", "needs-review", "", "", "Confirm the slide deck is intentionally public and approved."),
        ("media-prayer-card-image", "Prayer Card Image", "Google Drive", "needs-review", "", "", "Do not publish unless the image is intentionally public and cleared for use."),
    ]
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        style_body_row(ws, row_idx, 1, len(headers))
    add_dropdown(ws, "D2:D500", '=Lists!$C$2:$C$9')
    set_column_widths(ws, {"A": 24, "B": 28, "C": 18, "D": 22, "E": 18, "F": 16, "G": 52})
    ws.freeze_panes = "A2"


def build_lists_sheet(ws):
    ws["A1"] = "Media Types"
    ws["B1"] = "Category Slugs"
    ws["C1"] = "Copyright Status"
    ws["D1"] = "Status"
    ws["E1"] = "Featured"
    ws["F1"] = "Collection IDs"

    for idx, value in enumerate(MEDIA_TYPE_OPTIONS, start=2):
        ws[f"A{idx}"] = value
    for idx, row in enumerate(CATEGORY_ROWS, start=2):
        ws[f"B{idx}"] = row[2]
    for idx, value in enumerate(COPYRIGHT_OPTIONS, start=2):
        ws[f"C{idx}"] = value
    for idx, value in enumerate(STATUS_OPTIONS, start=2):
        ws[f"D{idx}"] = value
    for idx, value in enumerate(FEATURED_OPTIONS, start=2):
        ws[f"E{idx}"] = value
    for idx, row in enumerate(COLLECTION_ROWS, start=2):
        ws[f"F{idx}"] = row[0]


def apply_media_validations(ws):
    add_dropdown(ws, "F2:F1000", '=Lists!$A$2:$A$9')
    add_dropdown(ws, "G2:G1000", '=Lists!$B$2:$B$22')
    add_dropdown(ws, "X2:X1000", '=Lists!$C$2:$C$9')
    add_dropdown(ws, "Z2:Z1000", '=Lists!$E$2:$E$3')
    add_dropdown(ws, "AA2:AA1000", '=Lists!$F$2:$F$7')
    add_dropdown(ws, "AE2:AE1000", '=Lists!$D$2:$D$6')


def add_dropdown(ws, cell_range: str, formula: str):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.prompt = "Choose from the dropdown options."
    dv.promptTitle = "Daily Oratory"
    dv.error = "Please select one of the allowed values."
    dv.errorTitle = "Invalid selection"
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_header_row(ws, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell)


def style_header_cell(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_body_row(ws, row_idx: int, start_col: int, end_col: int):
    for col_idx in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = BODY_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER


def set_column_widths(ws, mapping: dict[str, float]):
    for col, width in mapping.items():
        ws.column_dimensions[col].width = width


if __name__ == "__main__":
    main()
